import json
from django.shortcuts import render, redirect
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from apify_client import ApifyClient
from openpyxl import Workbook
from io import BytesIO
from .models import Candidate
from .forms import ScrapeForm
from django.conf import settings


def home(request):
    error = None
    success = None

    if request.method == 'POST':
        form = ScrapeForm(request.POST)
        if form.is_valid():
            platform = form.cleaned_data['platform']
            keywords = form.cleaned_data['keywords']
            education = form.cleaned_data.get('education', '')
            location = form.cleaned_data.get('location', '')
            max_items = form.cleaned_data['max_items']

            client = ApifyClient(settings.APIFY_TOKEN)

            try:
                if platform == 'linkedin':
                    print(f"🔍 Starting LinkedIn Scrape | Keywords: {keywords} | Location: {location}")

                    # Step 1: Search Profiles
                    search_actor = 'harvestapi/linkedin-profile-search'   # Your search actor ID
                    search_input = {
                        "autoQuerySegmentation": False,

                        "currentJobTitles": [
                            keywords
                        ],

                        "locations": [
                            location
                        ],

                        "maxItems": max_items,

                        "recentlyChangedJobs": False,
                        "recentlyPostedOnLinkedIn": False,

                        "yearsAtCurrentCompanyIds": [
                            "2"
                        ],

                        "yearsOfExperienceIds": [
                            "2"
                        ]
                    }

                    if education:
                        search_input["schools"] = [education]

                    print("SEARCH INPUT:", search_input)

                    print("Step 1: Running Search Actor...")

                    search_run = client.actor(search_actor).call(
                        run_input=search_input
                    )


                    search_items = list(
                        client.dataset(
                            search_run["defaultDatasetId"]
                        ).iterate_items()
                    )

                    print("SEARCH ITEMS:", search_items[:2])

                    print(f"✅ Search returned {len(search_items)} items")
                    # Extract Profile URLs
                    profile_urls = []
                    for item in search_items:
                        print(item.keys())
                        url = item.get('linkedinUrl')
                        if url and 'linkedin.com/in/' in str(url):
                            profile_urls.append(url)

                    print(f"✅ Extracted {len(profile_urls)} LinkedIn Profile URLs")
               

                    if not profile_urls:
                        error = "No profile URLs found. Try different keywords or approve the search actor."
                    else:
                        # Step 2: Scrape Full Details
                        detail_actor = 'harvestapi/linkedin-profile-scraper'   # Your detail actor ID
                        detail_input = {
                            "profileScraperMode": "Profile details no email ($4 per 1k)",
                            "queries": profile_urls[:max_items]
                        }
                        print(f"Step 2: Scraping {len(detail_input['queries'])} profiles using detail actor...")
                        detail_run = client.actor(detail_actor).call(run_input=detail_input, timeout_secs=600)
                        
                        items = list(client.dataset(detail_run["defaultDatasetId"]).iterate_items())
                        print(items)
                       

                        print(f"✅ Detail Actor returned {len(items)} full profiles")

                        # Save to Database
                        saved = 0
                        for item in items:
                            print(f"item {item}")
                            Candidate.objects.create(
                                source='linkedin',
                                name=item.get('fullName') or item.get('name') or 'N/A',
                                mobile_no=item.get('mobileNumber') or item.get('phone'),
                                email=item.get('email'),
                                address=item.get('location'),
                                education=json.dumps(item.get('education', [])),
                                skills=keywords,
                                raw_data=item
                            )
                            saved += 1

                        success = f"🎉 Successfully saved {saved} LinkedIn profiles!"

                else:
                    error = "Naukri not implemented yet."

            except Exception as e:
                error = f"Error: {str(e)}"
                print("❌ EXCEPTION:", str(e))
                if "approve" in str(e).lower():
                    error = "Please approve both Actors in Apify Console."

    else:
        form = ScrapeForm()

    return render(request, 'scraper/home.html', {
        'form': form,
        'error': error,
        'success': success
    })
def candidate_list(request):
    candidates = Candidate.objects.all().order_by('-scraped_at')
    return render(request, 'scraper/candidate_list.html', {'candidates': candidates})


def view_resume(request, candidate_id):
    candidate = get_object_or_404(Candidate, id=candidate_id)

    # raw_data may already be dict or string
    data = candidate.raw_data or {}

    if isinstance(data, str):
        import json
        try:
            data = json.loads(data)
        except:
            data = {}

    # Safe location extraction
    location = ""
    if isinstance(data.get('location'), dict):
        location = data.get('location', {}).get('linkedinText', '')
    else:
        location = data.get('location', '')

    # Profile Image
    profile_picture = None
    if isinstance(data.get('profilePicture'), dict):
        profile_picture = data.get('profilePicture', {}).get('url')

    context = {
        'candidate': candidate,
        'name': f"{data.get('firstName', '')} {data.get('lastName', '')}",
        'headline': data.get('headline'),
        'location': location,
        'about': data.get('about'),
        'linkedin_url': data.get('linkedinUrl'),
        'connections': data.get('connectionsCount'),
        'followers': data.get('followerCount'),
        'profile_picture': profile_picture,

        # Sections
        'experience': data.get('experience', []),
        'education': data.get('education', []),
        'skills': data.get('skills', []),
        'certifications': data.get('certifications', []),
        'projects': data.get('projects', []),
        'languages': data.get('languages', []),
    }

    return render(request, 'scraper/resume.html', context)

# ====================== Keep these functions unchanged ======================
def export_excel(request):
    candidates = Candidate.objects.all().order_by('-scraped_at')
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"

    headers = ['Name', 'Mobile', 'Email', 'Address', 'Education', 'Skills', 'Source', 'Scraped At']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row, cand in enumerate(candidates, 2):
        ws.cell(row=row, column=1, value=cand.name)
        ws.cell(row=row, column=2, value=cand.mobile_no)
        ws.cell(row=row, column=3, value=cand.email)
        ws.cell(row=row, column=4, value=cand.address)
        ws.cell(row=row, column=5, value=cand.education)
        ws.cell(row=row, column=6, value=cand.skills)
        ws.cell(row=row, column=7, value=cand.source)
        ws.cell(row=row, column=8, value=cand.scraped_at.strftime('%Y-%m-%d %H:%M'))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=candidates.xlsx'
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.getvalue())
    return response




def clear_data(request):
    if request.method == 'POST':
        Candidate.objects.all().delete()
        return redirect('candidate_list')
    return render(request, 'scraper/confirm_clear.html')
